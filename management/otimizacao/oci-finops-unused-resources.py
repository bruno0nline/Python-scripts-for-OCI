import oci
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timezone
import logging
import sys

class OCI_FinOps_Report:
    def __init__(self):
        self.config = oci.config.from_file()
        self.identity_client = oci.identity.IdentityClient(self.config)
        self.tenancy_id = self.config["tenancy"]
        self.workbook = openpyxl.Workbook()

        self.sheets = {
            "Stopped Instances": ["Region", "Compartment", "Instance Name", "Instance OCID", "State", "Shape", "Created Time"],
            "Unattached Volumes": ["Region", "Compartment", "Volume Name", "Volume OCID", "Size (GB)", "State", "Created Time", "Last Backup Time"],
            "Unused Public IPs": ["Region", "Compartment", "Public IP", "OCID", "Assigned To", "State", "Created Time"],
            "Inactive DRGs & VPNs": ["Region", "Compartment", "Resource Name", "Type", "State", "Created Time"],
            "Unused Buckets": ["Region", "Compartment", "Bucket Name", "State", "Approximate Size (MB)", "Approximate Object Count"],
            "Unused File Systems": ["Region", "Compartment", "File System Name", "State", "Created Time"]
        }
        self._setup_sheets()

    def _setup_sheets(self):
        for sheet_name, headers in self.sheets.items():
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
            else:
                sheet = self.workbook.create_sheet(title=sheet_name)
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
            self.sheets[sheet_name] = sheet

        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])

    def _get_regions(self):
        logging.info("Buscando regiões ativas na tenancy...")
        regions = [r.region_name for r in self.identity_client.list_region_subscriptions(self.tenancy_id).data]
        logging.info(f"Regiões encontradas: {', '.join(regions)}")
        return regions

    def _get_compartments(self, tenancy_id):
        logging.info("Buscando compartimentos...")
        compartments = self.identity_client.list_compartments(
            tenancy_id, compartment_id_in_subtree=True
        ).data
        compartments.append(self.identity_client.get_compartment(tenancy_id).data)
        return compartments

    def collect_resources(self):
        regions = self._get_regions()
        
        for region in regions:
            logging.info(f"\n--- Coletando dados na região: {region} ---")
            self.config['region'] = region
            
            try:
                compute_client = oci.core.ComputeClient(self.config)
                blockstorage_client = oci.core.BlockstorageClient(self.config)
                network_client = oci.core.VirtualNetworkClient(self.config)
                object_storage_client = oci.object_storage.ObjectStorageClient(self.config)
                file_storage_client = oci.file_storage.FileStorageClient(self.config)
                load_balancer_client = oci.load_balancer.LoadBalancerClient(self.config)
            except Exception as e:
                logging.error(f"Erro ao inicializar clientes OCI na região {region}: {e}")
                continue

            # Obtém os domínios de disponibilidade específicos para a região atual
            try:
                availability_domains = self.identity_client.list_availability_domains(self.tenancy_id).data
            except Exception as e:
                logging.error(f"Erro ao buscar Domínios de Disponibilidade na região {region}: {e}")
                continue

            compartments = self._get_compartments(self.tenancy_id)
            for compartment in compartments:
                if compartment.lifecycle_state != "ACTIVE":
                    continue
                logging.info(f"  Verificando compartimento: {compartment.name}")

                # Stopped Instances
                try:
                    instances = compute_client.list_instances(compartment_id=compartment.id, lifecycle_state="STOPPED").data
                    for instance in instances:
                        self.sheets["Stopped Instances"].append([
                            region, compartment.name, instance.display_name, instance.id,
                            instance.lifecycle_state, instance.shape, instance.time_created.strftime('%Y-%m-%d %H:%M:%S')
                        ])
                except Exception as e:
                    logging.error(f"  Erro ao listar instâncias paradas em {compartment.name}: {e}")
                
                # Unattached Volumes
                try:
                    volumes = blockstorage_client.list_volumes(compartment_id=compartment.id).data
                    for volume in volumes:
                        attachments = compute_client.list_volume_attachments(compartment_id=compartment.id, volume_id=volume.id).data
                        if not attachments and volume.lifecycle_state == "AVAILABLE":
                             self.sheets["Unattached Volumes"].append([
                                region, compartment.name, volume.display_name, volume.id, volume.size_in_gbs,
                                volume.lifecycle_state, volume.time_created.strftime('%Y-%m-%d %H:%M:%S'), "N/A"
                            ])
                except Exception as e:
                    logging.error(f"  Erro ao listar volumes em {compartment.name}: {e}")
                
                # Unused Public IPs
                try:
                    public_ips = network_client.list_public_ips(scope="REGION", compartment_id=compartment.id).data
                    for ip in public_ips:
                        if ip.assigned_entity_id is None:
                            self.sheets["Unused Public IPs"].append([
                                region, compartment.name, ip.ip_address, ip.id, "Unassigned",
                                ip.lifecycle_state, ip.time_created.strftime('%Y-%m-%d %H:%M:%S')
                            ])
                except Exception as e:
                    logging.error(f"  Erro ao listar IPs públicos em {compartment.name}: {e}")

                # Inactive DRGs & VPNs
                try:
                    drgs = network_client.list_drgs(compartment_id=compartment.id).data
                    for drg in drgs:
                        if drg.lifecycle_state != "AVAILABLE":
                            self.sheets["Inactive DRGs & VPNs"].append([
                                region, compartment.name, drg.display_name, "DRG", drg.lifecycle_state, drg.time_created.strftime('%Y-%m-%d %H:%M:%S')
                            ])
                    
                    vpn_connections = network_client.list_ipsec_connections(compartment_id=compartment.id).data
                    for vpn in vpn_connections:
                        if vpn.lifecycle_state != "AVAILABLE":
                             self.sheets["Inactive DRGs & VPNs"].append([
                                region, compartment.name, vpn.display_name, "IPSec VPN", vpn.lifecycle_state, vpn.time_created.strftime('%Y-%m-%d %H:%M:%S')
                            ])
                except Exception as e:
                    logging.error(f"  Erro ao listar DRGs/VPNs em {compartment.name}: {e}")

                # Unused Object Storage Buckets
                try:
                    namespace = object_storage_client.get_namespace().data
                    buckets = object_storage_client.list_buckets(namespace, compartment_id=compartment.id).data
                    for bucket in buckets:
                        bucket_details = object_storage_client.get_bucket(namespace, bucket.name).data
                        if bucket_details.approximate_count == 0:
                            self.sheets["Unused Buckets"].append([
                                region, compartment.name, bucket.name, "Unused", bucket_details.approximate_size if bucket_details.approximate_size else 0,
                                bucket_details.approximate_count
                            ])
                except Exception as e:
                    logging.error(f"  Erro ao listar buckets em {compartment.name}: {e}")
                
                # Unused File Systems
                try:
                    for ad in availability_domains:
                        file_systems = file_storage_client.list_file_systems(compartment_id=compartment.id, availability_domain=ad.name).data
                        for fs in file_systems:
                            if fs.lifecycle_state != "AVAILABLE":
                                self.sheets["Unused File Systems"].append([
                                    region, compartment.name, fs.display_name, fs.lifecycle_state, fs.time_created.strftime('%Y-%m-%d %H:%M:%S')
                                ])
                except Exception as e:
                    logging.error(f"  Erro ao listar File Systems em {compartment.name}: {e}")
        
    def save_report(self):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"oci_finops_report_{timestamp}.xlsx"
            self.workbook.save(file_name)
            logging.info(f"\n✅ Relatório de otimização de custos salvo com sucesso: '{file_name}'")
        except Exception as e:
            logging.error(f"Erro ao salvar o relatório em Excel: {e}")

def main():
    report = OCI_FinOps_Report()
    report.collect_resources()
    report.save_report()

if __name__ == "__main__":
    main()