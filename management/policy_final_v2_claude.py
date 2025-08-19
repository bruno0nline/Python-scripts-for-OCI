#!/usr/bin/env python3
"""
OCI IAM Policy Audit Script

This script audits OCI IAM users, groups, policies, and dynamic groups,
generating a comprehensive Excel report with security analysis.

Requirements:
    - oci
    - openpyxl
    - OCI config file properly configured
"""

import os
import sys
import logging
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime
from collections import defaultdict, Counter

import oci
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(f'iam_audit_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
    ]
)
logger = logging.getLogger(__name__)


@dataclass
class SecurityRisk:
    """Data class for security risks."""
    policy_name: str
    risk_type: str
    details: str
    severity: str = "Medium"


class OCIIAMAuditor:
    """OCI IAM Auditor class for comprehensive IAM analysis."""
    
    def __init__(self, config_file: Optional[str] = None, profile: str = "DEFAULT"):
        """Initialize the auditor with OCI configuration."""
        try:
            # Use OCI SDK's default behavior when no config_file specified
            if config_file is None:
                self.config = oci.config.from_file(profile_name=profile)
                logger.info("📁 Using default OCI config location")
            else:
                self.config = oci.config.from_file(config_file, profile)
                logger.info(f"📁 Using OCI config at: {config_file}")
            
            self.tenancy_id = self.config["tenancy"]
            self.identity_client = oci.identity.IdentityClient(self.config)
            logger.info("✅ OCI client initialized successfully")
            logger.info(f"🏢 Using tenancy: {self.tenancy_id}")
            
        except oci.exceptions.ConfigFileNotFound as e:
            logger.error(f"❌ OCI config file not found: {e}")
            logger.error("💡 Please ensure your OCI config file exists at ~/.oci/config")
            raise
        except oci.exceptions.InvalidConfig as e:
            logger.error(f"❌ Invalid OCI configuration: {e}")
            logger.error("💡 Please check your OCI config file format and credentials")
            raise
        except Exception as e:
            logger.error(f"❌ Failed to initialize OCI client: {e}")
            raise
    
    @staticmethod
    def list_all_results(client_func, **kwargs) -> List:
        """Helper to handle pagination for any OCI list call."""
        try:
            return oci.pagination.list_call_get_all_results(client_func, **kwargs).data
        except Exception as e:
            logger.error(f"Error in pagination: {e}")
            return []
    
    def fetch_users(self) -> List:
        """Fetch all IAM users."""
        logger.info("🔍 Fetching users...")
        return self.list_all_results(
            self.identity_client.list_users, 
            compartment_id=self.tenancy_id
        )
    
    def fetch_groups(self) -> List:
        """Fetch all IAM groups."""
        logger.info("🔍 Fetching groups...")
        return self.list_all_results(
            self.identity_client.list_groups, 
            compartment_id=self.tenancy_id
        )
    
    def fetch_dynamic_groups(self) -> List:
        """Fetch all dynamic groups."""
        logger.info("🔍 Fetching dynamic groups...")
        return self.list_all_results(
            self.identity_client.list_dynamic_groups, 
            compartment_id=self.tenancy_id
        )
    
    def fetch_compartments(self) -> List:
        """Fetch all compartments including root."""
        logger.info("🔍 Fetching compartments...")
        compartments = self.list_all_results(
            self.identity_client.list_compartments,
            compartment_id=self.tenancy_id,
            compartment_id_in_subtree=True,
            access_level="ANY"
        )
        # Add root compartment
        root_compartment = oci.identity.models.Compartment(
            id=self.tenancy_id, 
            name="root"
        )
        compartments.append(root_compartment)
        return compartments
    
    def fetch_policies(self, compartments: List) -> List:
        """Fetch all policies from all compartments."""
        logger.info("🔍 Fetching policies from all compartments...")
        policies = []
        
        for compartment in compartments:
            try:
                comp_policies = self.list_all_results(
                    self.identity_client.list_policies, 
                    compartment_id=compartment.id
                )
                policies.extend(comp_policies)
            except Exception as e:
                logger.warning(f"Failed to fetch policies from compartment {compartment.name}: {e}")
        
        return policies
    
    def build_user_group_mapping(self, groups: List) -> Dict[str, List[str]]:
        """Build mapping of users to their groups."""
        logger.info("🔗 Building user-group mappings...")
        user_group_map = {}
        
        for group in groups:
            try:
                group_members = self.list_all_results(
                    self.identity_client.list_user_group_memberships,
                    compartment_id=self.tenancy_id,
                    group_id=group.id
                )
                for member in group_members:
                    user_group_map.setdefault(member.user_id, []).append(group.name)
            except Exception as e:
                logger.warning(f"Failed to fetch members for group {group.name}: {e}")
        
        return user_group_map
    
    def analyze_security_risks(self, policies: List) -> Tuple[List[SecurityRisk], Dict]:
        """Analyze policies for security risks."""
        logger.info("🔒 Analyzing security risks...")
        
        risks = []
        all_statements = []
        
        # High-risk patterns
        HIGH_RISK_PATTERNS = [
            "manage all-resources",
            "use all-resources", 
            "inspect tenancy",
            "manage tenancy",
            "manage iam",
            "manage authentication-policies"
        ]
        
        MEDIUM_RISK_PATTERNS = [
            "manage compute",
            "manage database",
            "manage object-storage-buckets",
            "manage load-balancers"
        ]
        
        for policy in policies:
            if not policy.statements:
                risks.append(SecurityRisk(
                    policy_name=policy.name,
                    risk_type="Orphaned Policy",
                    details="Policy has no statements",
                    severity="Low"
                ))
                continue
            
            for statement in policy.statements:
                all_statements.append((statement, policy.name))
                statement_lower = statement.lower()
                
                # Check for high-risk permissions
                for pattern in HIGH_RISK_PATTERNS:
                    if pattern in statement_lower:
                        risks.append(SecurityRisk(
                            policy_name=policy.name,
                            risk_type="High-Risk Permission",
                            details=f"Contains '{pattern}': {statement[:100]}...",
                            severity="High"
                        ))
                
                # Check for medium-risk permissions
                for pattern in MEDIUM_RISK_PATTERNS:
                    if pattern in statement_lower:
                        risks.append(SecurityRisk(
                            policy_name=policy.name,
                            risk_type="Broad Permission",
                            details=f"Contains '{pattern}': {statement[:100]}...",
                            severity="Medium"
                        ))
        
        # Analyze duplicates
        statement_counter = Counter([stmt for stmt, _ in all_statements])
        duplicate_statements = {
            stmt: count for stmt, count in statement_counter.items() 
            if count > 1
        }
        
        analysis = {
            'all_statements': all_statements,
            'duplicate_statements': duplicate_statements,
            'statement_counter': statement_counter
        }
        
        return risks, analysis
    
    def create_excel_report(self, data: Dict, filename: str = None) -> str:
        """Create comprehensive Excel report."""
        if not filename:
            filename = f"iam_policy_audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info("📊 Generating Excel report...")
        workbook = openpyxl.Workbook()
        
        # Create sheets
        self._create_summary_sheet(workbook, data)
        self._create_users_sheet(workbook, data['users'], data['user_group_map'])
        self._create_policies_sheet(workbook, data['policies'])
        self._create_dynamic_groups_sheet(workbook, data['dynamic_groups'])
        self._create_risks_sheet(workbook, data['risks'])
        self._create_duplicates_sheet(workbook, data['analysis']['duplicate_statements'], data['analysis'])
        
        # Remove default sheet and reorder
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Move summary to first position
        workbook._sheets.insert(0, workbook._sheets.pop(workbook._sheets.index(workbook['Summary'])))
        
        workbook.save(filename)
        logger.info(f"✅ Excel Report saved: {filename}")
        return filename
    
    def _create_summary_sheet(self, workbook, data):
        """Create summary sheet with key metrics."""
        sheet = workbook.create_sheet("Summary")
        
        # Headers
        headers = ["Metric", "Value", "Status"]
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        # Basic metrics
        metrics = [
            ("Total IAM Users", len(data['users']), "ℹ️"),
            ("Active Users", len([u for u in data['users'] if u.lifecycle_state == 'ACTIVE']), "✅"),
            ("Total Groups", len(data['groups']), "ℹ️"),
            ("Total Dynamic Groups", len(data['dynamic_groups']), "ℹ️"),
            ("Total IAM Policies", len(data['policies']), "ℹ️"),
            ("Total Policy Statements", len(data['analysis']['all_statements']), "ℹ️"),
            ("Duplicate Statements", len(data['analysis']['duplicate_statements']), 
             "⚠️" if data['analysis']['duplicate_statements'] else "✅"),
            ("High-Risk Policies", len([r for r in data['risks'] if r.severity == 'High']), 
             "🚨" if any(r.severity == 'High' for r in data['risks']) else "✅"),
            ("Total Security Risks", len(data['risks']), 
             "⚠️" if data['risks'] else "✅")
        ]
        
        for metric, value, status in metrics:
            sheet.append([metric, value, status])
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)
    
    def _create_users_sheet(self, workbook, users, user_group_map):
        """Create users sheet."""
        sheet = workbook.create_sheet("IAM Users")
        
        headers = ["User Name", "User OCID", "Status", "Groups", "Risk Level"]
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        for user in users:
            group_names = ", ".join(user_group_map.get(user.id, ["No Group"]))
            risk_level = "Low" if user.lifecycle_state == "ACTIVE" else "Medium"
            
            sheet.append([
                user.name,
                user.id,
                user.lifecycle_state,
                group_names,
                risk_level
            ])
        
        self._auto_adjust_columns(sheet)
    
    def _create_policies_sheet(self, workbook, policies):
        """Create policies sheet."""
        sheet = workbook.create_sheet("IAM Policies")
        
        headers = ["Policy Name", "Statement", "Compartment ID", "Created Time"]
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        for policy in policies:
            for statement in policy.statements:
                sheet.append([
                    policy.name,
                    statement,
                    policy.compartment_id,
                    policy.time_created.strftime('%Y-%m-%d %H:%M:%S') if policy.time_created else "Unknown"
                ])
        
        self._auto_adjust_columns(sheet)
    
    def _create_dynamic_groups_sheet(self, workbook, dynamic_groups):
        """Create dynamic groups sheet."""
        sheet = workbook.create_sheet("Dynamic Groups")
        
        headers = ["Name", "Description", "Matching Rule", "State"]
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        for dg in dynamic_groups:
            sheet.append([
                dg.name,
                dg.description or "",
                dg.matching_rule,
                dg.lifecycle_state
            ])
        
        self._auto_adjust_columns(sheet)
    
    def _create_risks_sheet(self, workbook, risks):
        """Create risks analysis sheet."""
        sheet = workbook.create_sheet("Security Risks")
        
        headers = ["Policy Name", "Risk Type", "Severity", "Details"]
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        # Sort by severity (High -> Medium -> Low)
        severity_order = {"High": 0, "Medium": 1, "Low": 2}
        sorted_risks = sorted(risks, key=lambda x: severity_order.get(x.severity, 3))
        
        for risk in sorted_risks:
            row = sheet.max_row + 1
            sheet.append([risk.policy_name, risk.risk_type, risk.severity, risk.details])
            
            # Color code by severity
            if risk.severity == "High":
                fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
            elif risk.severity == "Medium":
                fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
            else:
                fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            for col in range(1, 5):
                sheet.cell(row=row, column=col).fill = fill
        
        self._auto_adjust_columns(sheet)
    
    def _create_duplicates_sheet(self, workbook, duplicate_statements, analysis):
        """Create duplicate statements sheet."""
        sheet = workbook.create_sheet("Duplicate Statements")
        
        headers = ["Duplicate Statement", "Count", "Used In Policies", "Impact"]
        sheet.append(headers)
        self._style_header_row(sheet, 1)
        
        # Build mapping of statements to policies
        used_in_policies = defaultdict(set)
        for stmt, policy_name in analysis['all_statements']:
            if stmt in duplicate_statements:
                used_in_policies[stmt].add(policy_name)
        
        for stmt, count in sorted(duplicate_statements.items(), key=lambda x: x[1], reverse=True):
            impact = "High" if count > 5 else "Medium" if count > 2 else "Low"
            policies_list = ", ".join(sorted(used_in_policies[stmt]))
            
            sheet.append([stmt, count, policies_list, impact])
        
        self._auto_adjust_columns(sheet)
    
    def _style_header_row(self, sheet, row_num):
        """Style header row with bold font and fill."""
        for cell in sheet[row_num]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def run_audit(self) -> str:
        """Run complete IAM audit and generate report."""
        try:
            logger.info("🚀 Starting OCI IAM audit...")
            
            # Fetch all data
            users = self.fetch_users()
            groups = self.fetch_groups()
            dynamic_groups = self.fetch_dynamic_groups()
            compartments = self.fetch_compartments()
            policies = self.fetch_policies(compartments)
            
            # Build relationships
            user_group_map = self.build_user_group_mapping(groups)
            
            # Analyze security
            risks, analysis = self.analyze_security_risks(policies)
            
            # Prepare data for report
            data = {
                'users': users,
                'groups': groups,
                'dynamic_groups': dynamic_groups,
                'policies': policies,
                'compartments': compartments,
                'user_group_map': user_group_map,
                'risks': risks,
                'analysis': analysis
            }
            
            # Generate report
            filename = self.create_excel_report(data)
            
            logger.info("🎉 Audit completed successfully!")
            logger.info(f"📊 Summary: {len(users)} users, {len(groups)} groups, "
                       f"{len(policies)} policies, {len(risks)} risks identified")
            
            return filename
            
        except Exception as e:
            logger.error(f"❌ Audit failed: {e}")
            raise


def main():
    """Main function."""
    try:
        # You can specify config file and profile here
        auditor = OCIIAMAuditor()
        filename = auditor.run_audit()
        print(f"\n✅ Audit completed successfully!")
        print(f"📄 Report saved as: {filename}")
        
    except KeyboardInterrupt:
        logger.info("🛑 Audit interrupted by user")
        sys.exit(1)
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()