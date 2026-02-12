import oci
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import PieChart, BarChart, Reference
import sys
import logging

# Configuração de Logs
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', handlers=[
    logging.StreamHandler(sys.stdout)
])

# Carrega a configuração do OCI
config = oci.config.from_file("~/.oci/config")

# Inicializa os clientes da OCI
identity_client = oci.identity.IdentityClient(config)
virtual_network_client = oci.core.VirtualNetworkClient(config)
compute_client = oci.core.ComputeClient(config)
block_storage_client = oci.core.BlockstorageClient(config)
object_storage_client = oci.object_storage.ObjectStorageClient(config)
database_client = oci.database.DatabaseClient(config)
load_balancer_client = oci.load_balancer.LoadBalancerClient(config)
cloud_advisor_client = oci.optimizer.OptimizerClient(config)
cloud_guard_client = oci.cloud_guard.CloudGuardClient(config)

# Obtém o namespace da tenancy
try:
    namespace = object_storage_client.get_namespace().data
except oci.exceptions.ServiceError as e:
    logging.error(f"Erro ao obter o namespace: {e.message}. Verifique as permissões do usuário.")
    sys.exit(1)

# Obtém o OCID da tenancy
tenancy_id = config["tenancy"]

# Inicializa o armazenamento de resultados
resources = {}
findings = {}
cloud_advisor_recommendations = []
cloud_guard_findings = []

try:
    # Busca todos os compartimentos
    logging.info("Buscando todos os compartimentos...")
    compartments = oci.pagination.list_call_get_all_results(
        identity_client.list_compartments,
        tenancy_id,
        compartment_id_in_subtree=True,
        access_level="ANY"
    ).data
    compartments.append(oci.identity.models.Compartment(id=tenancy_id, name="Tenancy Root"))

    # Descobre recursos em cada compartimento
    for compartment in compartments:
        if compartment.lifecycle_state == "ACTIVE":
            logging.info(f"Descobrindo recursos no compartimento: {compartment.name}")
            resources[compartment.name] = {}
            findings[compartment.name] = []

            # Descobre VCNs
            vcn_response = oci.pagination.list_call_get_all_results(
                virtual_network_client.list_vcns,
                compartment_id=compartment.id
            ).data
            vcn_findings = []
            for vcn in vcn_response:
                resources[compartment.name].setdefault("VCNs", []).append({"name": vcn.display_name, "id": vcn.id})
                # Verificação de boas práticas: CIDR aberto
                if vcn.cidr_block == "0.0.0.0/0":
                    vcn_findings.append(f"VCN '{vcn.display_name}' tem um bloco CIDR aberto.")
            findings[compartment.name].extend(vcn_findings)

            # Descobre instâncias de computação
            instance_response = oci.pagination.list_call_get_all_results(
                compute_client.list_instances,
                compartment_id=compartment.id
            ).data
            instance_findings = []
            for instance in instance_response:
                resources[compartment.name].setdefault("Compute Instances", []).append({
                    "name": instance.display_name,
                    "id": instance.id
                })

                # Verifica se a instância está usando a imagem mais recente
                try:
                    image_details = compute_client.get_image(instance.image_id).data
                    if "platform" in image_details.operating_system and not image_details.is_latest:
                        instance_findings.append(f"Instância '{instance.display_name}' não está usando a imagem de plataforma mais recente.")
                except oci.exceptions.ServiceError:
                    pass

                # Verifica se a autenticação é baseada em chave SSH
                if not instance.metadata or "ssh_authorized_keys" not in instance.metadata:
                    instance_findings.append(f"Instância '{instance.display_name}' não tem autenticação baseada em chave SSH configurada.")

                # Verifica se o login por senha está desabilitado
                if instance.metadata and "disable_password_auth" not in instance.metadata:
                    instance_findings.append(f"Instância '{instance.display_name}' tem login por senha habilitado.")

                # Verifica se agentes de log estão configurados
                if "logging_agent" not in instance.metadata or instance.metadata.get("logging_agent") != "configured":
                    instance_findings.append(f"Instância '{instance.display_name}' não tem agentes de log configurados.")

                # Verifica se NSGs restringem portas desnecessárias
                vnics = compute_client.list_vnic_attachments(compartment_id=compartment.id, instance_id=instance.id).data
                for vnic_attachment in vnics:
                    vnic = virtual_network_client.get_vnic(vnic_attachment.vnic_id).data
                    nsgs = vnic.nsg_ids
                    for nsg_id in nsgs:
                        try:
                            nsg_rules = oci.pagination.list_call_get_all_results(
                                virtual_network_client.list_network_security_group_security_rules,
                                network_security_group_id=nsg_id
                            ).data
                            for rule in nsg_rules:
                                if rule.direction == "INGRESS" and rule.source == "0.0.0.0/0":
                                    instance_findings.append(f"NSG da instância '{instance.display_name}' permite entrada irrestrita.")
                        except oci.exceptions.ServiceError:
                            pass

            findings[compartment.name].extend(instance_findings)

            # Descobre Block Volumes
            volume_response = oci.pagination.list_call_get_all_results(
                block_storage_client.list_volumes,
                compartment_id=compartment.id
            ).data
            volume_findings = []
            for volume in volume_response:
                resources[compartment.name].setdefault("Block Volumes", []).append({
                    "name": volume.display_name,
                    "id": volume.id
                })
                # Verifica se o volume está anexado a alguma instância
                attachments = oci.pagination.list_call_get_all_results(
                    compute_client.list_volume_attachments,
                    compartment_id=compartment.id,
                    volume_id=volume.id
                ).data
                if not attachments:
                    volume_findings.append(f"Volume '{volume.display_name}' não está anexado a nenhuma instância.")
                # Verificação de boas práticas: política de backup
                if not volume.is_auto_tune_enabled:
                    volume_findings.append(f"Volume '{volume.display_name}' não tem auto-tune ativado.")
            findings[compartment.name].extend(volume_findings)

            # Descobre buckets do Object Storage
            bucket_response = oci.pagination.list_call_get_all_results(
                object_storage_client.list_buckets,
                namespace_name=namespace,
                compartment_id=compartment.id
            ).data
            bucket_findings = []
            for bucket in bucket_response:
                resources[compartment.name].setdefault("Buckets", []).append({"name": bucket.name})
                # Busca detalhes do bucket para verificar acesso público
                bucket_details = object_storage_client.get_bucket(
                    namespace_name=namespace,
                    bucket_name=bucket.name
                ).data
                # Verificação de boas práticas: acesso público
                if bucket_details.public_access_type != "NoPublicAccess":
                    bucket_findings.append(f"Bucket '{bucket.name}' permite acesso público.")
                # Descobre objetos nos buckets
                object_response = oci.pagination.list_call_get_all_results(
                    object_storage_client.list_objects,
                    namespace_name=namespace,
                    bucket_name=bucket.name
                ).data
                resources[compartment.name].setdefault("Bucket Objects", []).extend([
                    {"bucket_name": bucket.name, "object_name": obj.name} for obj in object_response.objects
                ])
            findings[compartment.name].extend(bucket_findings)

            # Descobre Autonomous Databases
            adb_response = oci.pagination.list_call_get_all_results(
                database_client.list_autonomous_databases,
                compartment_id=compartment.id
            ).data
            adb_findings = []
            for adb in adb_response:
                resources[compartment.name].setdefault("Autonomous Databases", []).append({
                    "name": adb.display_name,
                    "id": adb.id
                })
                # Verificação de boas práticas: tipo de carga de trabalho
                if adb.db_workload != "OLTP":
                    adb_findings.append(f"ADB '{adb.display_name}' não está otimizado para cargas de trabalho OLTP.")
            findings[compartment.name].extend(adb_findings)

            # Descobre Load Balancers
            lb_response = oci.pagination.list_call_get_all_results(
                load_balancer_client.list_load_balancers,
                compartment_id=compartment.id
            ).data
            lb_findings = []
            for lb in lb_response:
                resources[compartment.name].setdefault("Load Balancers", []).append({
                    "name": lb.display_name,
                    "id": lb.id
                })
                # Verificação de boas práticas: forma flexível
                if not lb.shape_name.startswith("flexible"):
                    lb_findings.append(f"Load Balancer '{lb.display_name}' não está usando uma forma flexível.")
            findings[compartment.name].extend(lb_findings)

    # Descobre recomendações do Cloud Advisor
    try:
        advisor_recommendations = oci.pagination.list_call_get_all_results(
            cloud_advisor_client.list_recommendations,
            compartment_id=tenancy_id,
            compartment_id_in_subtree=True
        ).data
        for recommendation in advisor_recommendations:
            cloud_advisor_recommendations.append({
                "Name": recommendation.name,
                "Recommendation": recommendation.description
            })
    except oci.exceptions.ServiceError as e:
        logging.error(f"Erro no serviço Cloud Advisor: {e.message}")

    # Descobre descobertas do Cloud Guard
    try:
        cloud_guard_problems = oci.pagination.list_call_get_all_results(
            cloud_guard_client.list_problems,
            compartment_id=tenancy_id,
            compartment_id_in_subtree=True
        ).data
        for problem in cloud_guard_problems:
            cloud_guard_findings.append({
                "Name": problem.resource_name,
                "Description": problem.labels
            })
    except oci.exceptions.ServiceError as e:
        logging.error(f"Erro no serviço Cloud Guard: {e.message}")

    # Exporta dados para JSON
    output_json = "oci_resources_audit.json"
    with open(output_json, "w") as file:
        json.dump({"resources": resources, "findings": findings, "cloud_advisor_recommendations": cloud_advisor_recommendations, "cloud_guard_findings": cloud_guard_findings}, file, indent=4)
    logging.info(f"Descoberta e validação de recursos concluídas. Resultados salvos em '{output_json}'.")

    # Exporta dados para Excel
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Sumário de Descobertas"

    # Adiciona sumário de descobertas
    summary_sheet.append(["Compartimento", "Problema", "Recomendação"])
    for compartment, issues in findings.items():
        for issue in issues:
            summary_sheet.append([compartment, issue, "Consulte as melhores práticas da OCI."])

    # Adiciona Recomendações do Cloud Advisor
    advisor_sheet = workbook.create_sheet(title="Cloud Advisor")
    advisor_sheet.append(["Nome", "Recomendação"])
    if cloud_advisor_recommendations:
        for recommendation in cloud_advisor_recommendations:
            advisor_sheet.append([recommendation["Name"], recommendation["Recommendation"]])
    else:
        advisor_sheet.append(["Nenhuma recomendação do Cloud Advisor encontrada."])

    # Adiciona Descobertas do Cloud Guard
    cloud_guard_sheet = workbook.create_sheet(title="Cloud Guard")
    cloud_guard_sheet.append(["Nome do Recurso", "Descrição"])
    if cloud_guard_findings:
        for finding in cloud_guard_findings:
            cloud_guard_sheet.append([finding["Name"], finding["Description"]])
    else:
        cloud_guard_sheet.append(["Nenhuma descoberta do Cloud Guard encontrada."])

    # Adiciona planilhas para cada tipo de recurso
    resource_types = ["VCNs", "Compute Instances", "Block Volumes", "Buckets", "Bucket Objects", "Autonomous Databases", "Load Balancers"]
    for resource_type in resource_types:
        sheet = workbook.create_sheet(title=resource_type)
        sheet.append(["Compartimento", "Nome", "ID"])
        for compartment, resource_data in resources.items():
            for item in resource_data.get(resource_type, []):
                sheet.append([compartment, item.get("name"), item.get("id", "N/A")])

    # Adiciona a planilha de visualização
    visualization_sheet = workbook.create_sheet(title="Visualizações")
    visualization_sheet.append(["Tipo de Recurso", "Contagem"])

    resource_counts = {rt: sum(len(data.get(rt, [])) for data in resources.values()) for rt in resource_types}

    for resource_type, count in resource_counts.items():
        visualization_sheet.append([resource_type, count])

    pie_chart = PieChart()
    pie_chart.title = "Distribuição de Recursos"
    pie_data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(resource_counts) + 1)
    pie_labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(resource_counts) + 1)
    pie_chart.add_data(pie_data, titles_from_data=False)
    pie_chart.set_categories(pie_labels)
    visualization_sheet.add_chart(pie_chart, "D2")

    bar_chart = BarChart()
    bar_chart.title = "Contagem de Recursos"
    bar_data = Reference(visualization_sheet, min_col=2, min_row=2, max_row=len(resource_counts) + 1)
    bar_labels = Reference(visualization_sheet, min_col=1, min_row=2, max_row=len(resource_counts) + 1)
    bar_chart.add_data(bar_data, titles_from_data=False)
    bar_chart.set_categories(bar_labels)
    visualization_sheet.add_chart(bar_chart, "D20")

    # Salva o arquivo Excel
    output_excel = "oci_resources_audit.xlsx"
    workbook.save(output_excel)
    logging.info(f"Detalhes e visualizações salvas em '{output_excel}'.")

except oci.exceptions.ServiceError as e:
    logging.error(f"Erro no serviço: {e}")
except Exception as e:
    logging.error(f"Ocorreu um erro inesperado: {e}")